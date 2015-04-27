#!/usr/bin/env python
# -*- coding: utf-8 -*-
# ctdailyqasumbmitter.py
"""Read and submit the CT Daily QA test results from an Excel file."""
# Copyright (c) 2015 Aditya Panchal

import resultssubmitter
import openpyxl
import pprint
import logging
logger = logging.getLogger('qatrackimport.ctdailyqasumbmitter')


class CTDailyQASubmitter(object):
    """Class that reads the CT Daily QA test results from an Excel (.xlsx) file
       and submits them to QATrack+"""
    def __init__(self, filename):

        self.filename = filename
        self.url = "http://127.0.0.1:8080/"
        self.username = 'admin'
        self.password = 'admin'

    def set_qatrack_server(self, url, username, password):
        """Setup the QA Track+ server settings."""

        self.url = url
        self.username = username
        self.password = password

    def read_excel_file(self):
        """Read the CT Daily QA Excel file from disk."""

        wb = openpyxl.load_workbook(filename=self.filename,
                                    use_iterators=True,
                                    data_only=True)
        self.ws = wb.get_active_sheet()

    def process_test(self, test, testnum, vartype):
        """Process the test result to make sure it is valid. Otherwise
           return a skipped test."""

        if test is None:
            return {"form-" + str(testnum) + "-skipped": "1"}
        else:
            return {"form-" + str(testnum) + "-value": vartype(test)}

    def convert_test_result(self, data, row):
        """Convert the test result into a dictonary compatible with the
           QATrack+ UnitTestCollection."""

        # If no sims were performed, return None
        if data[1] is None or "NO" in data[1].upper():
            logger.info("Skipping Row # %s (no data)", row)
            return None

        test_results = {
            "work_started": data[0].replace(hour=6).strftime("%d-%m-%Y %H:%M"),
            "work_completed": data[0].replace(
                hour=6, minute=30).strftime("%d-%m-%Y %H:%M"),
            "status": 2,  # Approved status
            "form-TOTAL_FORMS": "25",
            "form-INITIAL_FORMS": "25",
            "form-MAX_NUM_FORMS": "1000"
        }

        # Process tests 0-18 (numeric tests)
        for test in range(2, 21):
            test_results.update(self.process_test(data[test], test - 2, float))

        # Process tests 10 and 12 (multiple choice tests)
        if "form-10-value" in test_results:
            test_results["form-10-value"] = \
                int(test_results["form-10-value"] - 1)
        if "form-12-value" in test_results:
            test_results["form-12-value"] = \
                int(test_results["form-12-value"] - 2)

        # Process test 19 and 20 (laser / couch deviation tests)
        if not ((data[22] is None) or (data[21] is None)):
            test_results["form-19-value"] = \
                float(data[22]) * -1 if data[21] == 'R' else float(data[22])
        else:
            test_results["form-19-skipped"] = "1"

        # Process test 20 (laser / couch deviation test up/down)
        if not ((data[24] is None) or (data[23] is None)):
            test_results["form-20-value"] = \
                float(data[24]) * -1 if data[23] == 'P' else float(data[24])
        else:
            test_results["form-20-skipped"] = "1"

        # Process test 21 (couch test)
        test_results.update(self.process_test(data[25], 21, float))

        # Process tests 22-24 (boolean tests)
        for test in [26, 27, 28]:
            test_results.update(self.process_test(data[test], test - 4, str))
        for test in ["form-22-value", "form-23-value", "form-24-value"]:
            if test in test_results:
                test_results[test] = "1" \
                    if test_results[test].upper() == "X" else "0"

        # Operator initials
        comment = "" if data[29] is None else data[29]
        test_results["comment"] = "Performed by " + data[1] + \
            "\nRow " + str(row) + "\n" + comment

        logger.debug("Test Results: %s", pprint.pformat(test_results))

        return test_results

    def submit_data(self, startrow=None, endrow=None, utc=1,
                    progressfunc=None, updatefunc=None, dryrun=False):
        """Submit the test results to the QATrack+ server"""

        rs = resultssubmitter.ResultsSubmitter(
            self.url, self.username, self.password)

        # Row 55 is the first set of data with new procedure
        start = 55 if startrow is None else startrow
        end = self.ws.max_row if endrow is None else endrow

        # Iterate over the selected rows
        data_dimensions = 'B' + str(start) + ':AE' + str(end)
        dims = end - start
        rownum = start
        logger.info("Data dimensions: %s", data_dimensions)
        for row in self.ws.iter_rows(data_dimensions):
            logger.info("Reading Row # %s", rownum)
            data = [x.value for x in row]
            logger.debug("Data: %s %d", data, len(data))
            try:
                test_results = self.convert_test_result(data, rownum)
            except:
                if updatefunc:
                    updatefunc(utc, rownum)
                raise
                raise Exception("Error on row " + str(rownum) +
                                ". Please check data and retry.")
            else:
                # Update the progress function
                if progressfunc:
                    progressfunc("Reading record: " + str(rownum-start) +
                                 " of " + str(dims) + " [Row " +
                                 str(rownum) + "]")
                # If the test results aren't None, submit to server
                if test_results and not dryrun:
                    logger.info("Submitting Row # %s to server", rownum)
                    text = rs.submit_data(utc, test_results)
                    with open("result.html", 'w') as f:
                        f.write(text)

            rownum = rownum + 1
            # Update the update function after the result has been submitted
            if updatefunc:
                updatefunc(utc, rownum)

if __name__ == '__main__':

    import sys
    import argparse
    import logging
    import logging.handlers
    logger = logging.getLogger('qatrackimport')
    logger.setLevel(logging.INFO)
    ch = logging.StreamHandler()
    ch.setFormatter(logging.Formatter('%(levelname)s: %(message)s'))
    logger.addHandler(ch)

    # Set up argparser to parse the command-line arguments
    class DefaultParser(argparse.ArgumentParser):
        def error(self, message):
            sys.stderr.write('error: %s\n' % message)
            self.print_help()
            sys.exit(2)

    parser = DefaultParser(
        description="Read CT Daily QA test results from an Excel (.xls) " +
        "file and submit them to QATrack+.")
    parser.add_argument("filename",
                        help="Excel (.xlsx) file name")
    parser.add_argument("-s", "--startrow",
                        help="Starting row number",
                        type=int)
    parser.add_argument("-e", "--endrow",
                        help="Ending row number",
                        type=int)
    parser.add_argument("-d", "--debug",
                        help="Show debug log",
                        action="store_true")
    parser.add_argument("-y", "--dryrun",
                        help="Dry run mode (read data without submitting)",
                        action="store_true")

    # If there are no arguments, display help and exit
    if len(sys.argv) == 1:
        parser.print_help()
        sys.exit(1)
    args = parser.parse_args()

    # Set debug logging if the debug flag is set
    if args.debug:
        logger.setLevel(logging.DEBUG)

    # Read the CT Daily QA Excel file
    reader = CTDailyQASubmitter(args.filename)
    reader.read_excel_file()
    reader.submit_data(args.startrow, args.endrow, dryrun=args.dryrun)
